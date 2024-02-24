import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog
import datetime
# import open


import ctypes
import time
# MessageBox types
MB_OK = 0x0  # OK button
MB_OKCANCEL = 0x1  # OK and Cancel buttons
MB_YESNO = 0x4  # Yes and No buttons
MB_ICONINFORMATION = 0x40  # Information icon
MB_ICONWARNING = 0x30  # Warning icon
MB_ICONERROR = 0x10  # Error icon

# Function to display message box
def show_loading_alert(message):
    ctypes.windll.user32.MessageBoxW(0, message, "Loading", MB_OK )

# Function to display message box
def Finish_alert(message):
    ctypes.windll.user32.MessageBoxW(0, f"Successfully data stored in {message}", "Finished", MB_OK )








# Function to display message box
def show_message_box(title, message, style):
    return ctypes.windll.user32.MessageBoxW(0, message, title, style)

def show():
    result = show_message_box("Alert", "would you want to execute this file?", MB_OK | MB_ICONINFORMATION |MB_YESNO)
    if result == 7:  # 7 corresponds to No button
        show_loading_alert('Do you want to close?')
        exit()
        
def load():
    # Simulate some loading process
    show_loading_alert("Loading, please wait...")
    time.sleep(1)  # Simulating loading process
    


# Create a new workbook
wb = Workbook()

# Create a new worksheet within the workbook
ws = wb.active

def converter(data_list, reg,name):
    global first_table
    subject_grades = {'Name':name,'Reg No': reg}
    
    # Iterate over the list, skipping the first row (headers)
    for i in range(7, len(data_list), 7):
        subject_no = data_list[i + 1]
        grade = data_list[i + 6]
        subject_grades[subject_no] = grade
    
    print(subject_grades.values())
   
    if first_table:
        ws.append(list(subject_grades.keys()))
        first_table = False
        
    
    
    # Append the dictionary data to the worksheet
    ws.append(list(subject_grades.values()))
    
    

# Function to extract data from the HTML response
def extract_data(html_content, reg):
    # Parse the HTML content
    soup = BeautifulSoup(html_content, 'html.parser')
    table = soup.find_all('table')
    
    if len(table) > 2:
        name_table=table[0]
        mark_table = table[2]
        mark_data = []
        name_list=[]
        rows = mark_table.find_all('td')
        name_row=name_table.find_all('td')
        
        for detail in name_row:
            name_detail=detail.get_text(strip=True)
            name_list.append(name_detail)

        for cell in rows:
            cell_text = cell.get_text(strip=True)
            mark_data.append(cell_text)
        
        for item in name_list:
            # Check if the item starts with "Name of the Student :"
            if item.startswith('Name of the Student :'):
                # Extract the name
                name = item.split(':')[-1].strip()
                print(name)
                break  
        return converter(mark_data, reg,name)
    else:
        raise Exception("Mark table not found in HTML content.")

# URL and headers
url = "https://exam.pondiuni.edu.in/results/app.php?a=DisplayStudentResult"
headers = {
    'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Mobile Safari/537.36',
    'X-Requested-With': 'XMLHttpRequest',
    'Referer': 'https://exam.pondiuni.edu.in/results/result.php?r=21tk0055&e=C',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Accept-Language': 'en-GB,en-US;q=0.9,en;q=0.8,ta;q=0.7',
}


###create a Alert to run this program
show()



##we need to execute alert
# Create a Tkinter root window
root = tk.Tk()
root.withdraw()  # Hide the root window

# Prompt the user to select an Excel file using a file dialog
file_path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx *.xls")])

# Load the Excel file
workbook = load_workbook(file_path)
worksheet = workbook.active

# Initialize an empty list to store data from the first column
register_list = []
sem_list = []

# Iterate over each row and extract data from the first column
for row in worksheet.iter_rows(values_only=True):
    register_list.append(row[0])
    sem_list.append(row[1])

def switch(classer):
    inputer= classer.lower();
    if inputer =='first':
        return'A'
    elif inputer == 'second':
        return'B'
    elif inputer == 'third':
        return'C'
    elif inputer == 'fourth':
        return'D'
    elif inputer == 'fifth':
        return'E'
    elif inputer == 'sixth':
        return'F'
    
        
    


sem = switch(sem_list[0]);



first_table = True

# Initialize a list to store registration numbers for which errors occurred
error_reg_numbers = []

##loading alert
load()

# Main function to process requests for registration numbers
for reg_no in register_list:
    if reg_no is None:
        continue  # Skip None values
    data = {
        'r': reg_no,
        'e': sem
    }
    try:
        # Send the POST request
        response = requests.post(url, headers=headers, data=data, timeout=10)

        # Check if the request was successful (status code 200)
        if response.status_code == 200:
            # Extract data from HTML content
            html_content = response.json()['data']['html']
            
            # Print table data
            print("Table Data:")
            print("\n")
            
            extracted_data = extract_data(html_content, data['r'])
            
        else:
            print("Failed to retrieve the webpage. Status code:", response.status_code)
            # Store registration number for which error occurred
            error_reg_numbers.append(reg_no)
    except requests.Timeout:
        print("Request timed out for registration number:", reg_no)
        # Store registration number for which error occurred
        error_reg_numbers.append(reg_no)
    except requests.ConnectionError as e:
        print(f"Connection error occurred for registration number {reg_no}: {e}")
        # Store registration number for which error occurred
        error_reg_numbers.append(reg_no)
    except Exception as e:
        print(f"Error occurred for registration number {reg_no}: {e}")
        # Store registration number for which error occurred
        error_reg_numbers.append(reg_no)

# Process requests for registration numbers for which errors occurred
for reg_no in error_reg_numbers:
    data = {
        'r': reg_no,
        'e': sem
    }
    try:
        # Send the POST request
        response = requests.post(url, headers=headers, data=data, timeout=10)

        # Check if the request was successful (status code 200)
        if response.status_code == 200:
            # Extract data from HTML content
            html_content = response.json()['data']['html']
            
            # Print table data
            print("Table Data:")
            print("\n")
            
            extracted_data = extract_data(html_content, data['r'])
            
        else:
            print("Failed to retrieve the webpage. Status code:", response.status_code)
    except requests.Timeout:
        print("Request timed out for registration number:", reg_no)
    except requests.ConnectionError as e:
        print(f"Connection error occurred for registration number {reg_no}: {e}")
    except Exception as e:
        print(f"Error occurred for registration number {reg_no}: {e}")



# Save the workbook
filename=f"example_sorted_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
wb.save(filename)
Finish_alert(filename)
